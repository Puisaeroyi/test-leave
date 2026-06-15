"""Tests for approved leave export."""
from datetime import date
from decimal import Decimal
from io import BytesIO

from django.contrib.auth import get_user_model
from django.test import TestCase
from django.utils import timezone
from openpyxl import load_workbook
from rest_framework.test import APIClient

from leaves.models import LeaveCategory, LeaveRequest
from organizations.models import Entity

User = get_user_model()


class ExportApprovedLeavesTests(TestCase):
    def test_export_shows_both_approvers_and_notes(self):
        entity = Entity.objects.create(entity_name='Export Entity', code='EXPORT')
        employee = User.objects.create_user(
            email='employee-export@example.com',
            password='TestPass123!',
            role=User.Role.EMPLOYEE,
            employee_code='E-100',
            first_name='Export',
            last_name='Employee',
            entity=entity,
        )
        approver_1 = User.objects.create_user(
            email='approver-1-export@example.com',
            password='TestPass123!',
            role=User.Role.MANAGER,
            first_name='Approver',
            last_name='One',
        )
        approver_2 = User.objects.create_user(
            email='approver-2-export@example.com',
            password='TestPass123!',
            role=User.Role.MANAGER,
            first_name='Approver',
            last_name='Two',
        )
        hr = User.objects.create_user(
            email='hr-export@example.com',
            password='TestPass123!',
            role=User.Role.HR,
            entity=entity,
        )
        category = LeaveCategory.objects.create(
            category_name='Export Vacation',
            code='EXPORT_VACATION',
            balance_bucket='VACATION',
        )
        LeaveRequest.objects.create(
            user=employee,
            leave_category=category,
            start_date=date(2027, 2, 1),
            end_date=date(2027, 2, 1),
            shift_type='FULL_DAY',
            total_hours=Decimal('8.00'),
            reason='Export check',
            attachment_url='/media/leave_attachments/export-check.pdf',
            status='APPROVED',
            first_approver=approver_1,
            first_approval_status='APPROVED',
            first_approval_comment='ok-A',
            first_approval_at=timezone.now(),
            final_approver=approver_2,
            final_approval_status='APPROVED',
            final_approval_comment='ok-B',
            final_approval_at=timezone.now(),
            approved_by=approver_2,
            approved_at=timezone.now(),
            current_approval_step='COMPLETED',
        )

        client = APIClient()
        client.force_authenticate(user=hr)
        response = client.get('/api/v1/leaves/export/approved/?start_date=2027-02-01&end_date=2027-02-28')

        self.assertEqual(response.status_code, 200)
        workbook = load_workbook(BytesIO(response.content))
        sheet = workbook.active
        headers = [cell.value for cell in sheet[1]]
        self.assertIn('Approver 1', headers)
        self.assertIn('Approver 1 Status', headers)
        self.assertIn('Approver 1 Note', headers)
        self.assertIn('Approver 1 Date', headers)
        self.assertIn('Approver 2', headers)
        self.assertIn('Approver 2 Status', headers)
        self.assertIn('Approver 2 Note', headers)
        self.assertIn('Approver 2 Date', headers)
        self.assertIn('Approved By', headers)
        self.assertIn('Shift Type', headers)
        self.assertIn('Attachment URL', headers)
        self.assertIn('Created At', headers)
        self.assertIn('Updated At', headers)

        row = {headers[index]: sheet[2][index].value for index in range(len(headers))}
        self.assertEqual(row['Approver 1'], 'Approver One')
        self.assertEqual(row['Approver 1 Status'], 'APPROVED')
        self.assertEqual(row['Approver 1 Note'], 'ok-A')
        self.assertEqual(row['Approver 2'], 'Approver Two')
        self.assertEqual(row['Approver 2 Status'], 'APPROVED')
        self.assertEqual(row['Approver 2 Note'], 'ok-B')
        self.assertEqual(row['Approved By'], 'Approver Two')
        self.assertEqual(row['Shift Type'], 'Full Day')
        self.assertEqual(row['Attachment URL'], '/media/leave_attachments/export-check.pdf')
        self.assertTrue(row['Approver 1 Date'])
        self.assertTrue(row['Approver 2 Date'])
        self.assertTrue(row['Created At'])
        self.assertTrue(row['Updated At'])

    def test_export_includes_approved_leave_that_overlaps_date_range(self):
        entity = Entity.objects.create(entity_name='Overlap Export Entity', code='OVERLAP_EXPORT')
        employee = User.objects.create_user(
            email='overlap-export@example.com',
            password='TestPass123!',
            role=User.Role.EMPLOYEE,
            employee_code='E-200',
            first_name='Overlap',
            last_name='Employee',
            entity=entity,
        )
        hr = User.objects.create_user(
            email='hr-overlap-export@example.com',
            password='TestPass123!',
            role=User.Role.HR,
            entity=entity,
        )
        category = LeaveCategory.objects.create(
            category_name='Overlap Vacation',
            code='OVERLAP_VACATION',
            balance_bucket='VACATION',
        )
        LeaveRequest.objects.create(
            user=employee,
            leave_category=category,
            start_date=date(2027, 1, 31),
            end_date=date(2027, 2, 2),
            shift_type='FULL_DAY',
            total_hours=Decimal('24.00'),
            reason='Overlaps export window',
            status='APPROVED',
            approved_by=hr,
            approved_at=timezone.now(),
            current_approval_step='COMPLETED',
        )

        client = APIClient()
        client.force_authenticate(user=hr)
        response = client.get('/api/v1/leaves/export/approved/?start_date=2027-02-01&end_date=2027-02-28')

        self.assertEqual(response.status_code, 200)
        workbook = load_workbook(BytesIO(response.content))
        sheet = workbook.active
        headers = [cell.value for cell in sheet[1]]
        row = {headers[index]: sheet[2][index].value for index in range(len(headers))}
        self.assertEqual(row['Employee Code'], 'E-200')
        self.assertEqual(row['Reason'], 'Overlaps export window')

    def test_export_rejects_reversed_date_range(self):
        hr = User.objects.create_user(
            email='hr-reversed-export@example.com',
            password='TestPass123!',
            role=User.Role.HR,
        )
        client = APIClient()
        client.force_authenticate(user=hr)

        response = client.get('/api/v1/leaves/export/approved/?start_date=2027-03-01&end_date=2027-02-01')

        self.assertEqual(response.status_code, 400)
        self.assertIn(b'start_date must be before or equal to end_date', response.content)

    def test_hr_export_is_limited_to_own_entity_but_admin_export_is_global(self):
        own_entity = Entity.objects.create(entity_name='Own Export Entity', code='OWN_EXPORT')
        other_entity = Entity.objects.create(entity_name='Other Export Entity', code='OTHER_EXPORT')
        hr = User.objects.create_user(
            email='scoped-hr-export@example.com',
            password='TestPass123!',
            role=User.Role.HR,
            entity=own_entity,
        )
        admin = User.objects.create_user(
            email='global-admin-export@example.com',
            password='TestPass123!',
            role=User.Role.ADMIN,
        )
        own_employee = User.objects.create_user(
            email='own-employee-export@example.com',
            password='TestPass123!',
            role=User.Role.EMPLOYEE,
            entity=own_entity,
        )
        other_employee = User.objects.create_user(
            email='other-employee-export@example.com',
            password='TestPass123!',
            role=User.Role.EMPLOYEE,
            entity=other_entity,
        )
        category = LeaveCategory.objects.create(
            category_name='Scoped Export Vacation',
            code='SCOPED_EXPORT_VACATION',
            balance_bucket='VACATION',
        )
        for employee in (own_employee, other_employee):
            LeaveRequest.objects.create(
                user=employee,
                leave_category=category,
                start_date=date(2027, 5, 1),
                end_date=date(2027, 5, 1),
                shift_type='FULL_DAY',
                total_hours=Decimal('8.00'),
                status='APPROVED',
                approved_at=timezone.now(),
                current_approval_step='COMPLETED',
            )

        client = APIClient()
        client.force_authenticate(user=hr)
        hr_response = client.get(
            '/api/v1/leaves/export/approved/?start_date=2027-05-01&end_date=2027-05-31'
        )
        hr_sheet = load_workbook(BytesIO(hr_response.content)).active
        hr_emails = {row[2].value for row in hr_sheet.iter_rows(min_row=2)}
        self.assertEqual(hr_emails, {'own-employee-export@example.com'})

        client.force_authenticate(user=admin)
        admin_response = client.get(
            '/api/v1/leaves/export/approved/?start_date=2027-05-01&end_date=2027-05-31'
        )
        admin_sheet = load_workbook(BytesIO(admin_response.content)).active
        admin_emails = {row[2].value for row in admin_sheet.iter_rows(min_row=2)}
        self.assertEqual(
            admin_emails,
            {'own-employee-export@example.com', 'other-employee-export@example.com'},
        )
