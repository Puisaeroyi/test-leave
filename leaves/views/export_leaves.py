"""Export approved leave requests to Excel."""
from datetime import date
from io import BytesIO

from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from rest_framework.permissions import IsAuthenticated
from rest_framework.views import APIView

from leaves.models import LeaveRequest
from users.permissions import IsHROrAdmin


class ExportApprovedLeavesView(APIView):
    """Export approved leave requests within a date range to .xlsx."""

    permission_classes = [IsAuthenticated, IsHROrAdmin]

    def get(self, request):
        start_date = request.query_params.get("start_date")
        end_date = request.query_params.get("end_date")

        if not start_date or not end_date:
            return HttpResponse(
                "start_date and end_date query params required.",
                status=400,
            )

        try:
            start = date.fromisoformat(start_date)
            end = date.fromisoformat(end_date)
        except ValueError:
            return HttpResponse(
                "Invalid date format. Use YYYY-MM-DD.", status=400
            )

        if start > end:
            return HttpResponse(
                "start_date must be before or equal to end_date.",
                status=400,
            )

        # Build query filters - HR and ADMIN both have global access for export
        filters = {
            "status": "APPROVED",
            "start_date__lte": end,
            "end_date__gte": start,
        }

        queryset = (
            LeaveRequest.objects.filter(**filters)
            .select_related(
                "user",
                "leave_category",
                "approved_by",
                "user__entity",
                "user__location",
                "user__department",
                "user__approver_1",
                "user__approver_2",
                "first_approver",
                "final_approver",
            )
            .order_by("start_date", "user__last_name")
        )

        # Cap at 10,000 rows to prevent OOM
        MAX_EXPORT_ROWS = 10000
        total = queryset.count()
        if total > MAX_EXPORT_ROWS:
            return HttpResponse(
                f"Export too large ({total} rows). Max {MAX_EXPORT_ROWS}. Narrow the date range.",
                status=400,
            )
        queryset = queryset[:MAX_EXPORT_ROWS]

        wb = Workbook()
        ws = wb.active
        ws.title = "Approved Leaves"

        headers = [
            "Employee Code",
            "Employee Name",
            "Email",
            "Department",
            "Location",
            "Entity",
            "Leave Type",
            "Shift Type",
            "Start Date",
            "Start Time",
            "End Date",
            "End Time",
            "Total Hours",
            "Total Days",
            "Status",
            "Approver 1",
            "Approver 1 Status",
            "Approver 1 Note",
            "Approver 1 Date",
            "Approver 2",
            "Approver 2 Status",
            "Approver 2 Note",
            "Approver 2 Date",
            "Approved By",
            "Approved Date",
            "Reason",
            "Attachment URL",
            "Created At",
            "Updated At",
        ]

        # Header styling
        header_fill = PatternFill(
            start_color="4F81BD", end_color="4F81BD", fill_type="solid"
        )
        header_font = Font(color="FFFFFF", bold=True)
        header_align = Alignment(horizontal="center")

        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_align

        def display_name(person):
            if not person:
                return ""
            return f"{person.first_name} {person.last_name}".strip() or person.email

        def format_datetime(value):
            return value.strftime("%Y-%m-%d %H:%M") if value else ""

        def format_time(value):
            return value.strftime("%H:%M") if value else ""

        # Data rows
        for row_idx, lr in enumerate(queryset, 2):
            user = lr.user
            hours = float(lr.total_hours)
            row = [
                user.employee_code or "",
                display_name(user),
                user.email,
                getattr(user.department, "department_name", "") or "",
                getattr(user.location, "location_name", "") or "",
                getattr(user.entity, "entity_name", "") or "",
                getattr(lr.leave_category, "category_name", "") or "",
                lr.get_shift_type_display(),
                lr.start_date.isoformat(),
                format_time(lr.start_time),
                lr.end_date.isoformat(),
                format_time(lr.end_time),
                hours,
                round(hours / 8, 2),
                lr.status,
                display_name(lr.first_approver or user.approver_1),
                lr.first_approval_status,
                lr.first_approval_comment or "",
                format_datetime(lr.first_approval_at),
                display_name(lr.final_approver or user.approver_2),
                lr.final_approval_status,
                lr.final_approval_comment or "",
                format_datetime(lr.final_approval_at),
                display_name(lr.approved_by),
                format_datetime(lr.approved_at),
                lr.reason or "",
                lr.attachment_url or "",
                format_datetime(lr.created_at),
                format_datetime(lr.updated_at),
            ]
            for col_idx, value in enumerate(row, 1):
                ws.cell(row=row_idx, column=col_idx, value=value)

        # Auto-adjust column widths
        for col in ws.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                val = str(cell.value) if cell.value else ""
                max_len = max(max_len, len(val))
            ws.column_dimensions[col_letter].width = min(max_len + 2, 50)

        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)

        filename = f"approved_leaves_{start_date}_to_{end_date}.xlsx"
        response = HttpResponse(
            buf.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        return response
